"""Add a Submissions tab to Tax_Return.xlsx — single source of truth for what was filed."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

PATH = '/home/user/Claude/Tristan/projects/uk-rental/Tax_Return.xlsx'
wb = openpyxl.load_workbook(PATH)

name = "Submissions"
if name in wb.sheetnames:
    del wb[name]
ws = wb.create_sheet(name, 0)  # put first

TITLE = Font(bold=True, size=14, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="2E7D32")
SECTION = Font(bold=True, size=11, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="5B9BD5")
HDR = Font(bold=True, size=10)
HDR_FILL = PatternFill("solid", fgColor="E7E6E6")
BOLD = Font(bold=True)
ITALIC = Font(italic=True)
RESULT = PatternFill("solid", fgColor="FFE699")
GREEN = PatternFill("solid", fgColor="C6EFCE")
TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
TOTAL_FONT = Font(bold=True)

SEK = '#,##0'
GBP = '#,##0.00'

r = 1
ws.cell(row=r, column=1, value="Tax Returns 2025 — Submission Record").font = TITLE
ws.cell(row=r, column=1).fill = TITLE_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
ws.row_dimensions[r].height = 22
r += 1
ws.cell(row=r, column=1, value="Both returns submitted 2026-05-19. This tab is the single source of truth — all other tabs are working calcs.").font = ITALIC
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
r += 2

# Submission refs
ws.cell(row=r, column=1, value="Submission references").font = SECTION
ws.cell(row=r, column=1).fill = SECTION_FILL
for c in range(1, 6):
    ws.cell(row=r, column=c).fill = SECTION_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
r += 1

for h, val in [("Jurisdiction", "Value"), ("HMRC (UK)", ""), ("Skatteverket (Sweden)", "")]:
    pass

headers = ["Jurisdiction", "Form", "Date/time", "Reference", "Receipt file"]
for i, h in enumerate(headers, 1):
    ws.cell(row=r, column=i, value=h).font = HDR
    ws.cell(row=r, column=i).fill = HDR_FILL
r += 1

ws.cell(row=r, column=1, value="HMRC (UK)")
ws.cell(row=r, column=2, value="SA100 + SA105 2025/26")
ws.cell(row=r, column=3, value="2026-05-19 07:38 GMT")
ws.cell(row=r, column=4, value="2605755020")
ws.cell(row=r, column=5, value="UK_Tax_Return_2025-26_filed.pdf")
r += 1

ws.cell(row=r, column=1, value="Skatteverket (Sweden)")
ws.cell(row=r, column=2, value="INK1 + K4 avsnitt A 2025")
ws.cell(row=r, column=3, value="2026-05-19 10:14 CEST")
ws.cell(row=r, column=4, value="20260519101426199209225011262906")
ws.cell(row=r, column=5, value="Swedish_Tax_Return_2025_kvittens.pdf")
r += 2

# Swedish INK1 figures
ws.cell(row=r, column=1, value="Swedish INK1 2025 — submitted values").font = SECTION
ws.cell(row=r, column=1).fill = SECTION_FILL
for c in range(1, 6):
    ws.cell(row=r, column=c).fill = SECTION_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
r += 1

for i, h in enumerate(["Box", "Description", "SEK", "", "Working source"], 1):
    ws.cell(row=r, column=i, value=h).font = HDR
    ws.cell(row=r, column=i).fill = HDR_FILL
r += 1

ink1_rows = [
    ("1.1", "Lön, förmåner (förtryckt)", 705810, "Skatteverket förtryckt (KU)"),
    ("5.1", "Småhus/ägarlägenhet 0,75 % (Swedish property, sold 2026)", 671600, "Skatteverket förtryckt (taxeringsvärde)"),
    ("7.1", "Schablonintäkter (förtryckt)", 18360, "Förtryckt ISK/fond"),
    ("7.2", "Ränteinkomster (322 förtryckt + 2,730 HSBC UK)", 3052, "HSBC Interest tab"),
    ("7.3", "Överskott vid uthyrning av privatbostad", 310434, "Tax 25-26 tab; Knight Frank 2025"),
    ("7.4", "Vinst fondandelar (Vanguard ISA, K4)", 150141, "Capital Gains 2025 tab; K4-vanguard-working-2025.md"),
    ("8.1", "Ränteutgifter (Santander UK interest only)", 89876, "Santander tab; £6,955 × 12.92163"),
]
for box, desc, val, src in ink1_rows:
    ws.cell(row=r, column=1, value=box).font = BOLD
    ws.cell(row=r, column=2, value=desc)
    ws.cell(row=r, column=3, value=val).number_format = SEK
    ws.cell(row=r, column=5, value=src).font = ITALIC
    r += 1

r += 1
ws.cell(row=r, column=1, value="Avräkning utländsk skatt").font = BOLD
ws.cell(row=r, column=3, value=5608).number_format = SEK
ws.cell(row=r, column=3).fill = GREEN
ws.cell(row=r, column=5, value="9/12 UK 25/26 (£214) + 3/12 UK 24/25 (£1,092) × 12.92163").font = ITALIC
r += 2

# K4 detail
ws.cell(row=r, column=1, value="K4 avsnitt A — Vanguard ISA (4 funds)").font = SECTION
ws.cell(row=r, column=1).fill = SECTION_FILL
for c in range(1, 6):
    ws.cell(row=r, column=c).fill = SECTION_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
r += 1

for i, h in enumerate(["Antal", "Beteckning", "Försäljningspris (SEK)", "Omkostnadsbelopp (SEK)", "Vinst (SEK)"], 1):
    ws.cell(row=r, column=i, value=h).font = HDR
    ws.cell(row=r, column=i).fill = HDR_FILL
r += 1

k4_rows = [
    (74, "FTSE 100 Index Unit Trust Accumulation", 173014, 99683, 73331),
    (155, "S&P 500 UCITS ETF (VUSA)", 168353, 98869, 69484),
    (45, "FTSE 250 UCITS ETF (VMID)", 18869, 17495, 1374),
    (42, "FTSE Developed Europe ex UK (VERX)", 19957, 14005, 5952),
]
for ant, bet, fp, om, v in k4_rows:
    ws.cell(row=r, column=1, value=ant)
    ws.cell(row=r, column=2, value=bet)
    ws.cell(row=r, column=3, value=fp).number_format = SEK
    ws.cell(row=r, column=4, value=om).number_format = SEK
    ws.cell(row=r, column=5, value=v).number_format = SEK
    r += 1

ws.cell(row=r, column=2, value="Summa").font = TOTAL_FONT
ws.cell(row=r, column=3, value=380193).number_format = SEK
ws.cell(row=r, column=4, value=230052).number_format = SEK
ws.cell(row=r, column=5, value=150141).number_format = SEK
for c in range(1, 6):
    ws.cell(row=r, column=c).fill = TOTAL_FILL
    if ws.cell(row=r, column=c).value is not None:
        ws.cell(row=r, column=c).font = TOTAL_FONT
r += 2

# UK SA105 figures
ws.cell(row=r, column=1, value="UK SA100/SA105 2025/26 — submitted values").font = SECTION
ws.cell(row=r, column=1).fill = SECTION_FILL
for c in range(1, 6):
    ws.cell(row=r, column=c).fill = SECTION_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
r += 1

for i, h in enumerate(["Box", "Description", "GBP", "", "Source"], 1):
    ws.cell(row=r, column=i, value=h).font = HDR
    ws.cell(row=r, column=i).fill = HDR_FILL
r += 1

uk_rows = [
    ("SA100 TR3 box 2", "Untaxed UK interest (HSBC)", 242, "HSBC Interest tab"),
    ("SA105 box 20", "Total rents received (gross)", 32271, "76A Ingelow Road tab"),
    ("SA105 box 25", "Property repairs & maintenance", 7943, "76A Ingelow Road tab"),
    ("SA105 box 27", "Legal, management, prof fees", 2859, "Knight Frank"),
    ("SA105 box 38", "Adjusted profit", 21469, "Tax 25-26 tab"),
    ("SA105 box 40", "Taxable profit", 21469, "Tax 25-26 tab"),
    ("SA105 box 44", "Residential property finance costs", 7830, "Santander tab"),
    ("SA110 box 1", "Total tax due", 213.80, "Tax 25-26 tab"),
]
for box, desc, val, src in uk_rows:
    ws.cell(row=r, column=1, value=box).font = BOLD
    ws.cell(row=r, column=2, value=desc)
    ws.cell(row=r, column=3, value=val).number_format = GBP
    ws.cell(row=r, column=5, value=src).font = ITALIC
    r += 1

r += 1

# FX rates summary
ws.cell(row=r, column=1, value="FX rates used (Riksbanken)").font = SECTION
ws.cell(row=r, column=1).fill = SECTION_FILL
for c in range(1, 6):
    ws.cell(row=r, column=c).fill = SECTION_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
r += 1
for i, h in enumerate(["Period", "Rate (SEK/GBP)", "Use", "", ""], 1):
    ws.cell(row=r, column=i, value=h).font = HDR
    ws.cell(row=r, column=i).fill = HDR_FILL
r += 1
fx_summary = [
    ("Sep 2020 – Nov 2021 (monthly Medel)", "11.24–11.98", "Vanguard buys (per Capital Gains 2025 tab)"),
    ("Jun 2025 Medel", 12.94850, "Vanguard sale (all 4 funds)"),
    ("2025 årsmedel", 12.92163, "Recurring items (rental, interest, FTC)"),
]
for period, rate, use in fx_summary:
    ws.cell(row=r, column=1, value=period)
    if isinstance(rate, float):
        ws.cell(row=r, column=2, value=rate).number_format = '0.00000'
    else:
        ws.cell(row=r, column=2, value=rate)
    ws.cell(row=r, column=3, value=use).font = ITALIC
    r += 1

r += 1

# Companion files
ws.cell(row=r, column=1, value="Companion files (this folder)").font = SECTION
ws.cell(row=r, column=1).fill = SECTION_FILL
for c in range(1, 6):
    ws.cell(row=r, column=c).fill = SECTION_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
r += 1
files = [
    ("Tax_Filing_Guide_2025.md", "Canonical post-filing reference (positions + statutory anchors)"),
    ("Tax_Filing_Guide_2025.docx", "Word copy regenerated from .md"),
    ("K4-vanguard-working-2025.md", "Per-fund Vanguard genomsnittsmetoden detail"),
    ("Ovriga_upplysningar_2025.md", "Three blocks for Skatteverket free-text"),
    ("Swedish_Tax_Return_2025_filed.md", "Submission record incl. all values + kvittens"),
    ("Swedish_Tax_Return_2025_kvittens.pdf", "Skatteverket receipt PDF"),
    ("UK_Tax_Return_2025-26_filed.pdf", "HMRC submission PDF"),
    ("UK_Tax_Return_2025-26_submission_receipt.pdf", "HMRC receipt PDF"),
    ("vanguard-transactions/*.pdf", "Vanguard ISA transaction history (evidence)"),
    ("Next_Year_Checklist.md", "Roadmap for 2026 returns"),
    ("notes.md", "Mortgage history + tax-year totals"),
]
for f, desc in files:
    ws.cell(row=r, column=1, value=f).font = Font(name='Courier New', size=10)
    ws.cell(row=r, column=2, value=desc)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    r += 1

# Column widths
ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 50

ws.freeze_panes = "A4"

wb.save(PATH)
print(f"Submissions tab added to {PATH}")
print(f"Total sheets now: {wb.sheetnames}")
