"""Add Yield Analysis tab to Tax_Return.xlsx."""
import openpyxl
from openpyxl.styles import Font, PatternFill

PATH = '/home/user/Claude/Tristan/projects/uk-rental/Tax_Return.xlsx'
wb = openpyxl.load_workbook(PATH)

name = "Yield Analysis"
if name in wb.sheetnames:
    del wb[name]
ws = wb.create_sheet(name)

H = Font(bold=True, size=14, color="FFFFFF")
HF = PatternFill("solid", fgColor="4472C4")
S = Font(bold=True, size=11)
SF = PatternFill("solid", fgColor="D9E1F2")
R = PatternFill("solid", fgColor="FFE699")
G = PatternFill("solid", fgColor="C6EFCE")
Y = PatternFill("solid", fgColor="FFFFCC")
Bld = Font(bold=True)
It = Font(italic=True)

ws.cell(row=1, column=1, value="76a Ingelow Road — Investment Yield Analysis").font = H
ws.cell(row=1, column=1).fill = HF
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
ws.cell(row=2, column=1, value="Assuming gross-of-commission position for Swedish tax (£1,215 more conservative)").font = It

# === Parameters ===
r = 4
ws.cell(row=r, column=1, value="Inputs (edit to update yields)").font = S
for c in range(1, 6): ws.cell(row=r, column=c).fill = SF
r += 1
ws.cell(row=r, column=1, value="Property purchase price (2018)")
ws.cell(row=r, column=2, value=635000)
ws.cell(row=r, column=2).fill = Y
purchase_ref = f"$B${r}"
r += 1
ws.cell(row=r, column=1, value="Years held (to 2025)")
ws.cell(row=r, column=2, value=7)
ws.cell(row=r, column=2).fill = Y
years_ref = f"$B${r}"
r += 1
ws.cell(row=r, column=1, value="Current property value")
ws.cell(row=r, column=2, value=800000)
ws.cell(row=r, column=2).fill = Y
value_ref = f"$B${r}"
r += 1
ws.cell(row=r, column=1, value="Current mortgage outstanding (interest only)")
ws.cell(row=r, column=2, value=178764.94)
ws.cell(row=r, column=2).fill = Y
mortgage_ref = f"$B${r}"
r += 1
ws.cell(row=r, column=1, value="Equity in property (= value − mortgage)").font = Bld
ws.cell(row=r, column=2, value=f"={value_ref}-{mortgage_ref}").font = Bld
ws.cell(row=r, column=2).fill = R
equity_ref = f"$B${r}"
r += 1
ws.cell(row=r, column=1, value="Eventual CGT assumption (on current £165k appreciation)")
ws.cell(row=r, column=2, value=24000)
ws.cell(row=r, column=2).fill = Y
cgt_ref = f"$B${r}"
r += 2

# === CAGR / appreciation ===
ws.cell(row=r, column=1, value="Historical CAGR (from 2018 purchase)").font = S
for c in range(1, 6): ws.cell(row=r, column=c).fill = SF
r += 1
ws.cell(row=r, column=1, value="Total appreciation (current − purchase)")
ws.cell(row=r, column=2, value=f"={value_ref}-{purchase_ref}")
r += 1
ws.cell(row=r, column=1, value="Effective CGT rate on appreciation")
ws.cell(row=r, column=2, value=f"={cgt_ref}/({value_ref}-{purchase_ref})")
ws.cell(row=r, column=2).number_format = "0.00%"
cgt_rate_ref = f"$B${r}"
r += 1
ws.cell(row=r, column=1, value="CAGR (annualised)").font = Bld
ws.cell(row=r, column=2, value=f"=({value_ref}/{purchase_ref})^(1/{years_ref})-1").font = Bld
ws.cell(row=r, column=2).number_format = "0.00%"
ws.cell(row=r, column=2).fill = R
cagr_ref = f"$B${r}"
r += 1
ws.cell(row=r, column=1, value="Forward annual appreciation (at CAGR × current value)")
ws.cell(row=r, column=2, value=f"={value_ref}*{cagr_ref}")
fwd_appr_ref = f"$B${r}"
r += 1
ws.cell(row=r, column=1, value="Forward appreciation AFTER CGT")
ws.cell(row=r, column=2, value=f"={fwd_appr_ref}*(1-{cgt_rate_ref})")
fwd_appr_post_cgt_ref = f"$B${r}"
r += 2

# === Annual income statement 2025 (with gross-of-commission Swedish position) ===
ws.cell(row=r, column=1, value="Income statement 2025 (gross-of-commission Swedish, product fee in 2025)").font = S
for c in range(1, 6): ws.cell(row=r, column=c).fill = SF
r += 1
ws.cell(row=r, column=2, value="GBP").font = Bld
r += 1

# Rental income
ws.cell(row=r, column=1, value="Gross rental income (UK 25/26)")
ws.cell(row=r, column=2, value="='Tax 25-26'!$B$16")
rent_ref = f"$B${r}"
r += 1
ws.cell(row=r, column=1, value="Knight Frank commission")
ws.cell(row=r, column=2, value="='Tax 25-26'!$B$17")
comm_ref = f"$B${r}"
r += 1
ws.cell(row=r, column=1, value="Repairs & maintenance")
ws.cell(row=r, column=2, value="='Tax 25-26'!$B$19")
maint_ref = f"$B${r}"
r += 1
ws.cell(row=r, column=1, value="Mortgage interest")
ws.cell(row=r, column=2, value="=-'Tax 25-26'!$B$23")
mort_int_ref = f"$B${r}"
r += 1
ws.cell(row=r, column=1, value="Mortgage product fee (one-off, 2025)")
ws.cell(row=r, column=2, value="=-'Tax 25-26'!$B$24")
mort_fee_ref = f"$B${r}"
r += 1
ws.cell(row=r, column=1, value="UK tax on rental (net of finance credit)")
ws.cell(row=r, column=2, value="=-'Tax 25-26'!$B$32")
uk_tax_ref = f"$B${r}"
r += 1

# Swedish tax on rental property — compute marginally
# Swedish rental tax = (Box 7.3 with gross + adjusted - Box 8.1) * 30% - FTC
# For gross-of-commission scenario: Box 7.3 gross = $B$37 - 40k/fx - 20%*$B$37
# For simplicity, formula here directly:
ws.cell(row=r, column=1, value="Swedish tax on rental (gross-of-commission base, less FTC)")
# Box 7.3 gross = gross rent - 40k SEK - 20% gross = 0.80 × gross − 40k_in_GBP
# = 0.80 × $B$37 - $B$10/$B$4
# Box 8.1 = $B$51
# Net capital = above − Box 8.1
# Tax = 30% × that
# FTC = $B$56
# Net Swedish tax on rental = (0.80*B37 - 40k_in_GBP - B51)*30% - B56
ws.cell(row=r, column=2, value="=-((0.80*'Tax 25-26'!$B$37-'Tax 25-26'!$B$10/'Tax 25-26'!$B$4-'Tax 25-26'!$B$51)*0.30-'Tax 25-26'!$B$56)")
swed_tax_ref = f"$B${r}"
r += 1

# Net cash income
ws.cell(row=r, column=1, value="Net cash income after all costs and tax (2025)").font = Bld
ws.cell(row=r, column=2, value=f"={rent_ref}+{comm_ref}+{maint_ref}+{mort_int_ref}+{mort_fee_ref}+{uk_tax_ref}+{swed_tax_ref}").font = Bld
ws.cell(row=r, column=2).fill = G
net_cash_ref = f"$B${r}"
r += 2

# === Yield metrics ===
ws.cell(row=r, column=1, value="Yield metrics").font = S
for c in range(1, 6): ws.cell(row=r, column=c).fill = SF
r += 1
ws.cell(row=r, column=2, value="Value").font = Bld
ws.cell(row=r, column=3, value="% on current value").font = Bld
ws.cell(row=r, column=4, value="% on equity").font = Bld
r += 1

# Gross yield
ws.cell(row=r, column=1, value="Gross yield (rent / value)")
ws.cell(row=r, column=2, value=f"={rent_ref}")
ws.cell(row=r, column=3, value=f"=B{r}/{value_ref}")
ws.cell(row=r, column=3).number_format = "0.00%"
r += 1

# NOI yield (after operating costs, pre-leverage, pre-tax)
ws.cell(row=r, column=1, value="Net Operating Income (NOI — pre-leverage, pre-tax)")
ws.cell(row=r, column=2, value=f"={rent_ref}+{comm_ref}+{maint_ref}")
ws.cell(row=r, column=3, value=f"=B{r}/{value_ref}")
ws.cell(row=r, column=3).number_format = "0.00%"
r += 1

# Cash-on-cash (post-leverage, post-tax)
ws.cell(row=r, column=1, value="Net cash income (post all costs + tax, 2025)")
ws.cell(row=r, column=2, value=f"={net_cash_ref}")
ws.cell(row=r, column=3, value=f"=B{r}/{value_ref}")
ws.cell(row=r, column=3).number_format = "0.00%"
ws.cell(row=r, column=4, value=f"=B{r}/{equity_ref}")
ws.cell(row=r, column=4).number_format = "0.00%"
r += 2

# Steady-state (excluding one-off fee)
ws.cell(row=r, column=1, value="Steady-state cash income (excl. product fee)")
ws.cell(row=r, column=2, value=f"={net_cash_ref}-{mort_fee_ref}")
ws.cell(row=r, column=3, value=f"=B{r}/{value_ref}")
ws.cell(row=r, column=3).number_format = "0.00%"
ws.cell(row=r, column=4, value=f"=B{r}/{equity_ref}")
ws.cell(row=r, column=4).number_format = "0.00%"
steady_ref = f"$B${r}"
r += 2

# Total return (including appreciation)
ws.cell(row=r, column=1, value="Forward annual appreciation (CAGR × current value)")
ws.cell(row=r, column=2, value=f"={fwd_appr_ref}")
ws.cell(row=r, column=3, value=f"=B{r}/{value_ref}")
ws.cell(row=r, column=3).number_format = "0.00%"
ws.cell(row=r, column=4, value=f"=B{r}/{equity_ref}")
ws.cell(row=r, column=4).number_format = "0.00%"
r += 1

ws.cell(row=r, column=1, value="Total return — pre-CGT (steady cash + forward appreciation)").font = Bld
ws.cell(row=r, column=2, value=f"={steady_ref}+{fwd_appr_ref}").font = Bld
ws.cell(row=r, column=3, value=f"=B{r}/{value_ref}").font = Bld
ws.cell(row=r, column=3).number_format = "0.00%"
ws.cell(row=r, column=4, value=f"=B{r}/{equity_ref}").font = Bld
ws.cell(row=r, column=4).number_format = "0.00%"
ws.cell(row=r, column=2).fill = G
ws.cell(row=r, column=4).fill = G
r += 1

ws.cell(row=r, column=1, value="Total return — POST-CGT (steady cash + post-CGT appreciation)").font = Bld
ws.cell(row=r, column=2, value=f"={steady_ref}+{fwd_appr_post_cgt_ref}").font = Bld
ws.cell(row=r, column=3, value=f"=B{r}/{value_ref}").font = Bld
ws.cell(row=r, column=3).number_format = "0.00%"
ws.cell(row=r, column=4, value=f"=B{r}/{equity_ref}").font = Bld
ws.cell(row=r, column=4).number_format = "0.00%"
ws.cell(row=r, column=2).fill = G
ws.cell(row=r, column=4).fill = G
r += 2

# Comparison
ws.cell(row=r, column=1, value="Benchmarks (rough reference)").font = S
for c in range(1, 6): ws.cell(row=r, column=c).fill = SF
r += 1
ws.cell(row=r, column=1, value="UK 10-year gilts (risk-free)").font = It
ws.cell(row=r, column=3, value="~4-5%").font = It
r += 1
ws.cell(row=r, column=1, value="UK cash savings (instant access)").font = It
ws.cell(row=r, column=3, value="~4-5%").font = It
r += 1
ws.cell(row=r, column=1, value="UK FTSE 100 long-term nominal").font = It
ws.cell(row=r, column=3, value="~7-8%").font = It
r += 1
ws.cell(row=r, column=1, value="Global equity index (long-term)").font = It
ws.cell(row=r, column=3, value="~7-9%").font = It
r += 1
ws.cell(row=r, column=1, value="Your Vanguard ISA return (2020-2025)").font = It
ws.cell(row=r, column=3, value="~10%/yr").font = It
r += 2

# Notes
ws.cell(row=r, column=1, value="Notes").font = S
for c in range(1, 6): ws.cell(row=r, column=c).fill = SF
r += 1
for note in [
    "• 'Cash-on-cash' divides by equity, reflecting that you've only got ~£621k of own money in the flat.",
    "• 'NOI yield' is pre-leverage — useful for comparing flats to other property investments before mortgage effects.",
    "• Total return reflects assumption that 2018-2025 CAGR continues forward. London may not.",
    "• Post-CGT applies the implied 14.5% effective CGT rate (£24k on £165k gain) to forward appreciation.",
    "• Excludes: voids, refurbishment beyond annual maintenance, mortgage refinance risk after May 2027 (rate may rise from 4.38%).",
    "• Excludes: opportunity cost of equity. If £621k earned 5% in gilts = £31k/year passive — almost matches the steady-state cash yield.",
]:
    ws.cell(row=r, column=1, value=note).font = It
    r += 1

# Column widths
ws.column_dimensions['A'].width = 62
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 17
ws.column_dimensions['D'].width = 15

wb.save(PATH)
print("Saved Yield Analysis tab")

# Manual calc for chat
purchase = 635000
value = 800000
years = 7
mortgage = 178764.94
equity = value - mortgage
cgt = 24000
appreciation = value - purchase
cagr = (value / purchase) ** (1/years) - 1
fwd_appr = value * cagr
cgt_rate = cgt / appreciation
fwd_appr_post = fwd_appr * (1 - cgt_rate)

# Income
rent = 33900
comm = -5061.60
maint = -3243.99
mort_int = -6955.02
mort_fee = -1749
uk_tax = -214

# Swedish tax with gross base
gross_box73 = 0.80 * rent - 40000/12.92163  # GBP
box81 = 6955.02
swed_cap = (gross_box73 - box81) * 0.30
ftc = 1101.06 * 3/12 + 214 * 9/12  # using prior 24/25 tax + current 25/26
swed_tax = swed_cap - ftc

net_cash = rent + comm + maint + mort_int + mort_fee + uk_tax - swed_tax
steady = net_cash + 1749  # exclude one-off fee
noi = rent + comm + maint

print(f"\n=== KEY METRICS ===")
print(f"Property value: £{value:,}")
print(f"Equity in property: £{equity:,.0f}")
print(f"CAGR (2018-2025): {cagr*100:.2f}%")
print(f"Fwd annual appreciation @ CAGR: £{fwd_appr:,.0f}")
print(f"Net cash income 2025 (post all costs+tax, with one-off fee): £{net_cash:,.0f}")
print(f"Steady-state cash income (excl one-off fee): £{steady:,.0f}")
print(f"Gross yield: {rent/value*100:.2f}%")
print(f"NOI yield (pre-leverage): {noi/value*100:.2f}%")
print(f"Net cash yield on value: {net_cash/value*100:.2f}%")
print(f"Net cash yield on equity (CoC): {net_cash/equity*100:.2f}%")
print(f"Steady-state on equity: {steady/equity*100:.2f}%")
print(f"Total return pre-CGT (steady + appreciation):")
print(f"  on value: {(steady+fwd_appr)/value*100:.2f}%")
print(f"  on equity: {(steady+fwd_appr)/equity*100:.2f}%")
print(f"Total return POST-CGT:")
print(f"  on value: {(steady+fwd_appr_post)/value*100:.2f}%")
print(f"  on equity: {(steady+fwd_appr_post)/equity*100:.2f}%")
