"""Update HSBC Interest tab with Jan-Apr 2025 data — Swedish 2025 now complete."""
import openpyxl
from openpyxl.styles import Font, PatternFill

PATH = '/home/user/Claude/Tristan/projects/uk-rental/Tax_Return.xlsx'
wb = openpyxl.load_workbook(PATH)

name = "HSBC Interest"
if name in wb.sheetnames:
    del wb[name]
ws = wb.create_sheet(name)

H = Font(bold=True, size=14, color="FFFFFF")
HF = PatternFill("solid", fgColor="4472C4")
S = Font(bold=True, size=11)
SF = PatternFill("solid", fgColor="D9E1F2")
Bld = Font(bold=True)
It = Font(italic=True)

ws.cell(row=1, column=1, value="HSBC Savings Interest — by month credited").font = H
ws.cell(row=1, column=1).fill = HF
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

ws.cell(row=3, column=1, value="UK tax: report total in SA100 box 2 (Untaxed UK interest). PSA £1,000 covers basic rate taxpayers.").font = It
ws.cell(row=4, column=1, value="Swedish tax: report in Box 7.2 (Ränteinkomster) — taxed at 30% capital rate.").font = It

r = 6
headers = ["Credit date", "Period", "GROSS INTEREST (£)", "ADDED GROSS INTEREST (£)", "Total (£)", "UK Tax Year", "Swedish Year"]
for i, h in enumerate(headers, 1):
    ws.cell(row=r, column=i, value=h).font = Bld
    ws.cell(row=r, column=i).fill = SF

# Note: 01 Apr 25 credit is BEFORE 6 Apr 25 → falls in UK 24/25
# All other 2025 credits (May 25 onwards) are in UK 25/26
# 01 Apr 26 credit is BEFORE 6 Apr 26 → falls in UK 25/26
data = [
    ("01 Jan 25", "TO 31DEC2024", 4.82, 0,    "24/25", "2025"),
    ("01 Feb 25", "TO 31JAN2025", 4.58, 0,    "24/25", "2025"),
    ("01 Mar 25", "TO 28FEB2025", 1.95, 0,    "24/25", "2025"),
    ("01 Apr 25", "TO 31MAR2025", 1.96, 3.26, "24/25", "2025"),
    ("01 May 25", "TO 30APR2025", 3.12, 5.59, "25/26", "2025"),
    ("01 Jun 25", "TO 31MAY2025", 4.02, 7.84, "25/26", "2025"),
    ("01 Jul 25", "TO 30JUN2025", 11.04, 21.55, "25/26", "2025"),
    ("01 Aug 25", "TO 31JUL2025", 11.30, 21.75, "25/26", "2025"),
    ("01 Sep 25", "TO 31AUG2025", 10.00, 0,    "25/26", "2025"),
    ("01 Oct 25", "TO 30SEP2025", 9.65,  18.10, "25/26", "2025"),
    ("01 Nov 25", "TO 31OCT2025", 9.56,  18.47, "25/26", "2025"),
    ("01 Dec 25", "TO 30NOV2025", 14.08, 28.62, "25/26", "2025"),
    ("01 Jan 26", "TO 31DEC2025", 11.28, 0,    "25/26", "2026"),
    ("01 Feb 26", "TO 31JAN2026", 8.49,  0,    "25/26", "2026"),
    ("01 Mar 26", "TO 28FEB2026", 7.25,  0,    "25/26", "2026"),
    ("01 Apr 26", "TO 31MAR2026", 6.43,  13.65, "25/26", "2026"),
]

first_data_row = r + 1
for row in data:
    r += 1
    date, period, gross, added, uk_yr, swed_yr = row
    ws.cell(row=r, column=1, value=date)
    ws.cell(row=r, column=2, value=period)
    ws.cell(row=r, column=3, value=gross)
    ws.cell(row=r, column=4, value=added)
    ws.cell(row=r, column=5, value=f"=SUM(C{r}:D{r})")
    ws.cell(row=r, column=6, value=uk_yr)
    ws.cell(row=r, column=7, value=swed_yr)
last_data_row = r

r += 2
ws.cell(row=r, column=1, value="Totals by tax year").font = S
for c in range(1, 8): ws.cell(row=r, column=c).fill = SF
r += 1
ws.cell(row=r, column=2, value="Total (£)").font = Bld
ws.cell(row=r, column=3, value="Coverage").font = Bld
r += 1

ws.cell(row=r, column=1, value="UK 24/25 (partial — Jan-Apr 2025 credits)")
ws.cell(row=r, column=2, value=f"=SUMIF(F{first_data_row}:F{last_data_row},\"24/25\",E{first_data_row}:E{last_data_row})")
ws.cell(row=r, column=3, value="Partial — pre-Jan 2025 not loaded (prior year, not material)")
r += 1

ws.cell(row=r, column=1, value="UK 25/26 (full 12 months)").font = Bld
ws.cell(row=r, column=2, value=f"=SUMIF(F{first_data_row}:F{last_data_row},\"25/26\",E{first_data_row}:E{last_data_row})").font = Bld
ws.cell(row=r, column=3, value="Full year ✓")
r += 1

ws.cell(row=r, column=1, value="Swedish 2025 (full year)").font = Bld
ws.cell(row=r, column=2, value=f"=SUMIF(G{first_data_row}:G{last_data_row},\"2025\",E{first_data_row}:E{last_data_row})").font = Bld
ws.cell(row=r, column=3, value="Full year ✓")
r += 1

ws.cell(row=r, column=1, value="Swedish 2026 partial (Jan-Apr 2026)")
ws.cell(row=r, column=2, value=f"=SUMIF(G{first_data_row}:G{last_data_row},\"2026\",E{first_data_row}:E{last_data_row})")
ws.cell(row=r, column=3, value="Will need more data later").font = It
r += 2

ws.cell(row=r, column=1, value="Notes:").font = S
r += 1
for note in [
    "• Each month has a GROSS INTEREST line and (sometimes) an ADDED GROSS INTEREST line — both credit the account.",
    "• 01 Apr 25 credit falls in UK 24/25 (before 6 April cutoff); 01 Apr 26 credit falls in UK 25/26.",
    "• UK 25/26 PSA covers £1,000; with ~£242 interest, no UK tax due — but report on SA100 box 2.",
    "• Sweden taxes credited interest at 30% capital rate (Box 7.2 on INK1).",
]:
    ws.cell(row=r, column=1, value=note).font = It
    r += 1

ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 16
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 24
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 12

wb.save(PATH)

# Totals
uk2526 = sum((d[2] or 0) + (d[3] or 0) for d in data if d[4] == "25/26")
swed25 = sum((d[2] or 0) + (d[3] or 0) for d in data if d[5] == "2025")
uk2425 = sum((d[2] or 0) + (d[3] or 0) for d in data if d[4] == "24/25")
print(f"UK 24/25 partial (Jan-Apr 2025 only): £{uk2425:.2f}")
print(f"UK 25/26 FULL (12 months): £{uk2526:.2f}")
print(f"Swedish 2025 FULL (12 months): £{swed25:.2f}")
