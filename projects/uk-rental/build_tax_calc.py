"""Build tax calc sheets with FORMULAS referencing the data sheets.
User can edit inputs and outputs recalc automatically.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

SRC = '/home/user/Claude/projects/uk-rental/Tax_Return.xlsx'
DST = '/home/user/Claude/projects/uk-rental/Tax_Return.xlsx'

wb = openpyxl.load_workbook(SRC)
ING = '76A Ingelow Road'
SAN = 'Santander'

# Confirm sheet structure
# 76A Ingelow Road: Row 1 = Swedish year, Row 2 = UK tax year (float), Row 6 = income,
#                   Row 10 = mgmt commission, Row 11 = letting commission, Row 13 = maintenance
# Santander: Row 3 = UK tax year (str like "24/25"), Row 4 = Swedish year (str like "2024"),
#            Row 6 = interest, Row 8 = product fee

# Helper to find max col in a sheet
def max_col(ws):
    return ws.max_column

ing_max = max_col(wb[ING])
san_max = max_col(wb[SAN])

# Column letters
def col_range(sheet_name, row, last_col):
    return f"'{sheet_name}'!{get_column_letter(2)}{row}:{get_column_letter(last_col)}{row}"

ING_UK_ROW = f"'{ING}'!$C$2:${get_column_letter(ing_max)}$2"
ING_SW_ROW = f"'{ING}'!$C$1:${get_column_letter(ing_max)}$1"
ING_INCOME = f"'{ING}'!$C$6:${get_column_letter(ing_max)}$6"
ING_MGMT = f"'{ING}'!$C$10:${get_column_letter(ing_max)}$10"
ING_LETTING = f"'{ING}'!$C$11:${get_column_letter(ing_max)}$11"
ING_MAINT = f"'{ING}'!$C$13:${get_column_letter(ing_max)}$13"

SAN_UK_ROW = f"'{SAN}'!$B$3:${get_column_letter(san_max)}$3"
SAN_SW_ROW = f"'{SAN}'!$B$4:${get_column_letter(san_max)}$4"
SAN_INT = f"'{SAN}'!$B$6:${get_column_letter(san_max)}$6"
SAN_FEE = f"'{SAN}'!$B$8:${get_column_letter(san_max)}$8"

def build_sheet(wb, name, uk_float_val, uk_str_val, swed_yr,
                fx_value, fx_label, uk_prior_label_cell, uk_prior_tax_cell):
    """Build a tax calc sheet with formulas."""
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)

    # Styles
    H = Font(bold=True, size=14, color="FFFFFF")
    HF = PatternFill("solid", fgColor="4472C4")
    S = Font(bold=True, size=11)
    SF = PatternFill("solid", fgColor="D9E1F2")
    R = PatternFill("solid", fgColor="FFE699")
    G = PatternFill("solid", fgColor="C6EFCE")
    Bld = Font(bold=True)
    It = Font(italic=True)

    r = 1
    ws.cell(row=r, column=1, value=f"Tax calculation — UK {uk_str_val} + Swedish {swed_yr}").font = H
    ws.cell(row=r, column=1).fill = HF
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 2

    # === PARAMETERS (editable) ===
    ws.cell(row=r, column=1, value="Parameters (edit these to recalc):").font = S
    for c in range(1, 6): ws.cell(row=r, column=c).fill = SF
    r += 1

    # FX
    ws.cell(row=r, column=1, value="FX rate (SEK per GBP)").font = Bld
    ws.cell(row=r, column=2, value=fx_value)
    ws.cell(row=r, column=3, value=fx_label)
    fx_ref = f"$B${r}"
    r += 1
    # UK tax year (label & numeric)
    ws.cell(row=r, column=1, value="UK tax year (label/float)")
    ws.cell(row=r, column=2, value=uk_float_val)
    ws.cell(row=r, column=3, value=uk_str_val)
    uk_float_ref = f"$B${r}"
    uk_str_ref = f"$C${r}"
    r += 1
    # Swedish year
    ws.cell(row=r, column=1, value="Swedish tax year")
    ws.cell(row=r, column=2, value=swed_yr)
    ws.cell(row=r, column=3, value=str(swed_yr))
    swed_num_ref = f"$B${r}"
    swed_str_ref = f"$C${r}"
    r += 1
    # UK personal allowance
    ws.cell(row=r, column=1, value="UK personal allowance")
    ws.cell(row=r, column=2, value=12570)
    pa_ref = f"$B${r}"
    r += 1
    # UK basic rate
    ws.cell(row=r, column=1, value="UK basic rate")
    ws.cell(row=r, column=2, value=0.20)
    rate_ref = f"$B${r}"
    r += 1
    # UK financing credit rate
    ws.cell(row=r, column=1, value="UK financing-cost credit rate")
    ws.cell(row=r, column=2, value=0.20)
    fc_rate_ref = f"$B${r}"
    r += 1
    # Sweden std deduction
    ws.cell(row=r, column=1, value="Swedish standard deduction (SEK)")
    ws.cell(row=r, column=2, value=40000)
    se_std_ref = f"$B${r}"
    r += 1
    # Sweden % deduction
    ws.cell(row=r, column=1, value="Swedish % deduction (of rental income)")
    ws.cell(row=r, column=2, value=0.20)
    se_pct_ref = f"$B${r}"
    r += 1
    # Sweden capital rate
    ws.cell(row=r, column=1, value="Swedish capital income tax rate")
    ws.cell(row=r, column=2, value=0.30)
    se_cap_ref = f"$B${r}"
    r += 2

    # === UK SECTION ===
    ws.cell(row=r, column=1, value=f"UK tax year {uk_str_val} (6 April to 5 April)").font = S
    for c in range(1, 6): ws.cell(row=r, column=c).fill = SF
    r += 1
    ws.cell(row=r, column=2, value="GBP").font = Bld
    r += 1

    # Rental Income
    ws.cell(row=r, column=1, value="Rental Income")
    ws.cell(row=r, column=2, value=f"=SUMIF({ING_UK_ROW},{uk_float_ref},{ING_INCOME})")
    income_ref = f"$B${r}"
    r += 1

    # Commission (mgmt + letting)
    ws.cell(row=r, column=1, value="Knight Frank Commission (mgmt + letting)")
    ws.cell(row=r, column=2, value=f"=SUMIF({ING_UK_ROW},{uk_float_ref},{ING_MGMT})+SUMIF({ING_UK_ROW},{uk_float_ref},{ING_LETTING})")
    comm_ref = f"$B${r}"
    r += 1

    # Net before expenses
    ws.cell(row=r, column=1, value="Net Income before Expenses").font = Bld
    ws.cell(row=r, column=2, value=f"={income_ref}+{comm_ref}").font = Bld
    net_pre_ref = f"$B${r}"
    r += 1

    # Maintenance
    ws.cell(row=r, column=1, value="Repair and Maintenance Expenses")
    ws.cell(row=r, column=2, value=f"=SUMIF({ING_UK_ROW},{uk_float_ref},{ING_MAINT})")
    maint_ref = f"$B${r}"
    r += 1

    # Taxable profit
    ws.cell(row=r, column=1, value="Taxable Property Profit (SA105 box 38/40)").font = Bld
    ws.cell(row=r, column=2, value=f"={net_pre_ref}+{maint_ref}").font = Bld
    ws.cell(row=r, column=2).fill = R
    uk_profit_ref = f"$B${r}"
    r += 2

    # Memo: finance costs
    ws.cell(row=r, column=1, value="Memo — residential property finance costs (SA105 box 44):").font = It
    r += 1
    ws.cell(row=r, column=1, value="  Mortgage interest")
    ws.cell(row=r, column=2, value=f"=SUMIF({SAN_UK_ROW},{uk_str_ref},{SAN_INT})")
    uk_int_ref = f"$B${r}"
    r += 1
    ws.cell(row=r, column=1, value="  Mortgage product fees")
    ws.cell(row=r, column=2, value=f"=SUMIF({SAN_UK_ROW},{uk_str_ref},{SAN_FEE})")
    uk_fee_ref = f"$B${r}"
    r += 1
    ws.cell(row=r, column=1, value="  Total residential finance costs").font = Bld
    ws.cell(row=r, column=2, value=f"={uk_int_ref}+{uk_fee_ref}").font = Bld
    uk_fin_ref = f"$B${r}"
    r += 2

    # Tax calc
    ws.cell(row=r, column=1, value="UK tax calculation").font = S
    for c in range(1, 6): ws.cell(row=r, column=c).fill = SF
    r += 1
    ws.cell(row=r, column=1, value="Total taxable income (property profit)")
    ws.cell(row=r, column=2, value=f"={uk_profit_ref}")
    r += 1
    ws.cell(row=r, column=1, value="Above personal allowance")
    ws.cell(row=r, column=2, value=f"=MAX(0,{uk_profit_ref}-{pa_ref})")
    above_pa_ref = f"$B${r}"
    r += 1
    ws.cell(row=r, column=1, value="Basic rate tax (= above PA × basic rate)")
    ws.cell(row=r, column=2, value=f"={above_pa_ref}*{rate_ref}")
    basic_tax_ref = f"$B${r}"
    r += 1
    ws.cell(row=r, column=1, value="Less: financing cost credit (finance costs × credit rate)")
    ws.cell(row=r, column=2, value=f"=-{uk_fin_ref}*{fc_rate_ref}")
    fin_credit_ref = f"$B${r}"
    r += 1
    ws.cell(row=r, column=1, value="Net UK Tax").font = Bld
    ws.cell(row=r, column=2, value=f"={basic_tax_ref}+{fin_credit_ref}").font = Bld
    ws.cell(row=r, column=2).fill = G
    uk_tax_ref = f"$B${r}"
    r += 3

    # === SWEDISH SECTION ===
    ws.cell(row=r, column=1, value=f"Swedish tax year {swed_yr} (1 January to 31 December)").font = S
    for c in range(1, 6): ws.cell(row=r, column=c).fill = SF
    r += 1
    ws.cell(row=r, column=2, value="GBP").font = Bld
    ws.cell(row=r, column=3, value="SEK").font = Bld
    r += 1

    # Rental Income
    ws.cell(row=r, column=1, value="Rental Income")
    ws.cell(row=r, column=2, value=f"=SUMIF({ING_SW_ROW},{swed_num_ref},{ING_INCOME})")
    ws.cell(row=r, column=3, value=f"=B{r}*{fx_ref}")
    sw_income_ref = f"$B${r}"
    r += 1

    # Commission
    ws.cell(row=r, column=1, value="Knight Frank Commission")
    ws.cell(row=r, column=2, value=f"=SUMIF({ING_SW_ROW},{swed_num_ref},{ING_MGMT})+SUMIF({ING_SW_ROW},{swed_num_ref},{ING_LETTING})")
    ws.cell(row=r, column=3, value=f"=B{r}*{fx_ref}")
    sw_comm_ref = f"$B${r}"
    r += 1

    # Base
    ws.cell(row=r, column=1, value="Rental income net of commission (basis for schablonavdrag)").font = Bld
    ws.cell(row=r, column=2, value=f"={sw_income_ref}+{sw_comm_ref}").font = Bld
    ws.cell(row=r, column=3, value=f"=B{r}*{fx_ref}").font = Bld
    ws.cell(row=r, column=2).fill = R
    ws.cell(row=r, column=3).fill = R
    sw_base_ref = f"$B${r}"
    r += 2

    # Schablonavdrag
    ws.cell(row=r, column=1, value="Schablonavdrag (replaces actual expense deduction):").font = It
    r += 1
    ws.cell(row=r, column=1, value="  Standard SEK 40,000 deduction")
    ws.cell(row=r, column=2, value=f"=-{se_std_ref}/{fx_ref}")
    ws.cell(row=r, column=3, value=f"=-{se_std_ref}")
    sw_std_ref = f"$B${r}"
    r += 1
    ws.cell(row=r, column=1, value="  20% of rental income (net of commission)")
    ws.cell(row=r, column=2, value=f"=-{sw_base_ref}*{se_pct_ref}")
    ws.cell(row=r, column=3, value=f"=B{r}*{fx_ref}")
    sw_pct_ref = f"$B${r}"
    r += 1

    # Box 7.3
    ws.cell(row=r, column=1, value="Box 7.3 — Överskott vid uthyrning av privatbostad").font = Bld
    ws.cell(row=r, column=2, value=f"={sw_base_ref}+{sw_std_ref}+{sw_pct_ref}").font = Bld
    ws.cell(row=r, column=3, value=f"=B{r}*{fx_ref}").font = Bld
    ws.cell(row=r, column=2).fill = G
    ws.cell(row=r, column=3).fill = G
    sw_box73_ref = f"$B${r}"
    sw_box73_sek_ref = f"$C${r}"
    r += 2

    # Memo repairs not deductible
    ws.cell(row=r, column=1, value="Memo — repairs (not deductible under schablonavdrag):").font = It
    ws.cell(row=r, column=2, value=f"=SUMIF({ING_SW_ROW},{swed_num_ref},{ING_MAINT})").font = It
    r += 2

    # Box 8.1
    ws.cell(row=r, column=1, value="Box 8.1 — Ränteutgifter (mortgage interest + fees)").font = S
    for c in range(1, 6): ws.cell(row=r, column=c).fill = SF
    r += 1
    ws.cell(row=r, column=1, value="Mortgage interest (calendar year)")
    ws.cell(row=r, column=2, value=f"=SUMIF({SAN_SW_ROW},{swed_str_ref},{SAN_INT})")
    ws.cell(row=r, column=3, value=f"=B{r}*{fx_ref}")
    sw_int_ref = f"$B${r}"
    r += 1
    ws.cell(row=r, column=1, value="Mortgage product fees (deductible)")
    ws.cell(row=r, column=2, value=f"=SUMIF({SAN_SW_ROW},{swed_str_ref},{SAN_FEE})")
    ws.cell(row=r, column=3, value=f"=B{r}*{fx_ref}")
    sw_fee_ref = f"$B${r}"
    r += 1
    ws.cell(row=r, column=1, value="Box 8.1 total").font = Bld
    ws.cell(row=r, column=2, value=f"={sw_int_ref}+{sw_fee_ref}").font = Bld
    ws.cell(row=r, column=3, value=f"=B{r}*{fx_ref}").font = Bld
    ws.cell(row=r, column=2).fill = G
    ws.cell(row=r, column=3).fill = G
    sw_box81_ref = f"$B${r}"
    sw_box81_sek_ref = f"$C${r}"
    r += 2

    # Foreign tax credit
    ws.cell(row=r, column=1, value="Foreign tax credit (Avräkning utländsk skatt)").font = S
    for c in range(1, 6): ws.cell(row=r, column=c).fill = SF
    r += 1
    ws.cell(row=r, column=1, value=f"UK prior year tax × 3/12 (Jan-Mar fell in old UK year)")
    ws.cell(row=r, column=2, value=f"={uk_prior_tax_cell}*3/12")
    ws.cell(row=r, column=3, value=f"=B{r}*{fx_ref}")
    r += 1
    ws.cell(row=r, column=1, value=f"UK current year tax × 9/12 (Apr-Dec)")
    ws.cell(row=r, column=2, value=f"={uk_tax_ref}*9/12")
    ws.cell(row=r, column=3, value=f"=B{r}*{fx_ref}")
    r += 1
    ws.cell(row=r, column=1, value=f"Apportioned FTC for Swedish {swed_yr}").font = Bld
    ws.cell(row=r, column=2, value=f"={uk_prior_tax_cell}*3/12+{uk_tax_ref}*9/12").font = Bld
    ws.cell(row=r, column=3, value=f"=B{r}*{fx_ref}").font = Bld
    ws.cell(row=r, column=2).fill = G
    ws.cell(row=r, column=3).fill = G
    ftc_ref = f"$B${r}"
    ftc_sek_ref = f"$C${r}"
    r += 3

    # === Swedish tax to pay (marginal on rental property) ===
    ws.cell(row=r, column=1, value="Swedish tax payable on rental property (marginal calc)").font = S
    for c in range(1, 6): ws.cell(row=r, column=c).fill = SF
    r += 1
    ws.cell(row=r, column=1, value="Note: capital income (rental, interest, dividends) is taxed at flat 30% in Sweden.").font = It
    r += 1
    ws.cell(row=r, column=1, value="Other income (salary ~800,000 SEK) is in earned income (~50% marginal); does not affect this calc.").font = It
    r += 2

    ws.cell(row=r, column=1, value="Net capital income (Box 7.3 - Box 8.1)")
    ws.cell(row=r, column=2, value=f"={sw_box73_ref}-{sw_box81_ref}")
    ws.cell(row=r, column=3, value=f"=B{r}*{fx_ref}")
    net_cap_ref = f"$B${r}"
    net_cap_sek_ref = f"$C${r}"
    r += 1
    ws.cell(row=r, column=1, value="Capital tax @ 30% (gross, before foreign tax credit)")
    ws.cell(row=r, column=2, value=f"={net_cap_ref}*{se_cap_ref}")
    ws.cell(row=r, column=3, value=f"=B{r}*{fx_ref}")
    cap_tax_ref = f"$B${r}"
    r += 1
    ws.cell(row=r, column=1, value="Less: Foreign tax credit")
    ws.cell(row=r, column=2, value=f"=-{ftc_ref}")
    ws.cell(row=r, column=3, value=f"=B{r}*{fx_ref}")
    r += 1
    ws.cell(row=r, column=1, value="Swedish tax payable on this property").font = Bld
    ws.cell(row=r, column=2, value=f"={cap_tax_ref}-{ftc_ref}").font = Bld
    ws.cell(row=r, column=3, value=f"=B{r}*{fx_ref}").font = Bld
    ws.cell(row=r, column=2).fill = G
    ws.cell(row=r, column=3).fill = G
    r += 2

    # Column widths
    ws.column_dimensions['A'].width = 62
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30

    # Return key cell refs for cross-sheet linking
    return {
        "uk_tax_cell": f"'{name}'!{uk_tax_ref}",
        "ftc_cell": f"'{name}'!{ftc_ref}",
    }

# Build 24-25 first (its UK tax is referenced by 25-26 FTC calc)
# For 24-25 we need a "prior year UK tax" — use the filed £1,128 as a hardcoded value
r2425 = build_sheet(
    wb, "Tax 24-25",
    uk_float_val=24.25, uk_str_val="24/25", swed_yr=2024,
    fx_value=13.50453, fx_label="Riksbanken 2024 annual average",
    uk_prior_label_cell=None,
    uk_prior_tax_cell="1128",  # hardcoded prior UK 23/24 filed tax
)

# Build 25-26, referencing the 24-25 sheet's UK tax cell for the FTC apportionment
r2526 = build_sheet(
    wb, "Tax 25-26",
    uk_float_val=25.26, uk_str_val="25/26", swed_yr=2025,
    fx_value=12.92163, fx_label="Riksbanken 2025 annual average",
    uk_prior_label_cell=None,
    uk_prior_tax_cell=r2425["uk_tax_cell"],
)

wb.save(DST)
print(f"Saved {DST}")
