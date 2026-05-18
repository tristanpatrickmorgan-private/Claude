"""Add HSBC Interest tab to track UK savings interest by month."""
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

PATH = '/home/user/Claude/projects/uk-rental/Tax_Return.xlsx'
wb = openpyxl.load_workbook(PATH)

name = "HSBC Interest"
if name in wb.sheetnames:
    del wb[name]
ws = wb.create_sheet(name)

H = Font(bold=True, size=14, color="FFFFFF")
HF = PatternFill("solid", fgColor="4472C4")
S = Font(bold=True, size=11)
SF = PatternFill("solid", fgColor="D9E1F2")
G = PatternFill("solid", fgColor="C6EFCE")
M = PatternFill("solid", fgColor="FFCCCC")  # missing
Bld = Font(bold=True)
It = Font(italic=True)

# Title
ws.cell(row=1, column=1, value="HSBC Savings Interest — by month credited").font = H
ws.cell(row=1, column=1).fill = HF
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

ws.cell(row=3, column=1, value="UK tax: report total in SA100 box 2 (Untaxed UK interest). PSA £1,000 covers basic rate taxpayers.").font = It
ws.cell(row=4, column=1, value="Swedish tax: report in Box 7.2 (Ränteinkomster) — taxed at 30% capital rate.").font = It

# Header
r = 6
headers = ["Credit date", "Period", "GROSS INTEREST (£)", "ADDED GROSS INTEREST (£)", "Total (£)", "UK Tax Year", "Swedish Year"]
for i, h in enumerate(headers, 1):
    ws.cell(row=r, column=i, value=h).font = Bld
    ws.cell(row=r, column=i).fill = SF

# Data extracted from HSBC statements (date credited, period covered, GROSS, ADDED)
# UK tax year boundary: credits on/after 6 April are in new year; before 6 April are in old year
# Credit date 01 May 25 = UK 25/26; 01 Apr 26 = UK 25/26 (before 6 Apr 26); 01 May 26 = UK 26/27
data = [
    # Date,  period covered,  GROSS,  ADDED,  UK_year,  Swedish_year
    # Missing: Jan, Feb, Mar, Apr 2025 credits (need 2024-2025 statements)
    ("01 Jan 25", "TO 31DEC2024", None, None, "24/25", "2025", "MISSING"),
    ("01 Feb 25", "TO 31JAN2025", None, None, "24/25", "2025", "MISSING"),
    ("01 Mar 25", "TO 28FEB2025", None, None, "24/25", "2025", "MISSING"),
    ("01 Apr 25", "TO 31MAR2025", None, None, "24/25", "2025", "MISSING"),
    ("01 May 25", "TO 30APR2025", 3.12, 5.59, "25/26", "2025", "from statement"),
    ("01 Jun 25", "TO 31MAY2025", 4.02, 7.84, "25/26", "2025", "from statement"),
    ("01 Jul 25", "TO 30JUN2025", 11.04, 21.55, "25/26", "2025", "from statement"),
    ("01 Aug 25", "TO 31JUL2025", 11.30, 21.75, "25/26", "2025", "from statement"),
    ("01 Sep 25", "TO 31AUG2025", 10.00, 0,    "25/26", "2025", "from statement"),
    ("01 Oct 25", "TO 30SEP2025", 9.65,  18.10, "25/26", "2025", "from statement"),
    ("01 Nov 25", "TO 31OCT2025", 9.56,  18.47, "25/26", "2025", "from statement"),
    ("01 Dec 25", "TO 30NOV2025", 14.08, 28.62, "25/26", "2025", "from statement"),
    ("01 Jan 26", "TO 31DEC2025", 11.28, 0,    "25/26", "2026", "from statement"),
    ("01 Feb 26", "TO 31JAN2026", 8.49,  0,    "25/26", "2026", "from statement"),
    ("01 Mar 26", "TO 28FEB2026", 7.25,  0,    "25/26", "2026", "from statement"),
    ("01 Apr 26", "TO 31MAR2026", 6.43,  13.65, "25/26", "2026", "from statement"),
]

first_data_row = r + 1
for row in data:
    r += 1
    date, period, gross, added, uk_yr, swed_yr, source = row
    ws.cell(row=r, column=1, value=date)
    ws.cell(row=r, column=2, value=period)
    ws.cell(row=r, column=3, value=gross)
    ws.cell(row=r, column=4, value=added)
    if gross is not None or added is not None:
        ws.cell(row=r, column=5, value=f"=SUM(C{r}:D{r})")
    ws.cell(row=r, column=6, value=uk_yr)
    ws.cell(row=r, column=7, value=swed_yr)
    if source == "MISSING":
        for c in range(1, 8):
            ws.cell(row=r, column=c).fill = M
last_data_row = r

# Totals section
r += 2
ws.cell(row=r, column=1, value="Totals by tax year").font = S
for c in range(1, 8): ws.cell(row=r, column=c).fill = SF
r += 1
ws.cell(row=r, column=1, value="").font = Bld
ws.cell(row=r, column=2, value="Total (£)").font = Bld
ws.cell(row=r, column=3, value="Note").font = Bld
r += 1

# UK 25/26
ws.cell(row=r, column=1, value="UK 25/26 (credits 6 Apr 2025 to 5 Apr 2026)")
ws.cell(row=r, column=2, value=f"=SUMIF(F{first_data_row}:F{last_data_row},\"25/26\",E{first_data_row}:E{last_data_row})")
ws.cell(row=r, column=3, value="Full 12 months ✓")
uk2526_ref = f"$B${r}"
r += 1

# Swedish 2025
ws.cell(row=r, column=1, value="Swedish 2025 (credits Jan-Dec 2025)")
ws.cell(row=r, column=2, value=f"=SUMIF(G{first_data_row}:G{last_data_row},\"2025\",E{first_data_row}:E{last_data_row})")
ws.cell(row=r, column=3, value="MISSING Jan-Apr 2025 — partial").font = Font(color="990000")
swed2025_ref = f"$B${r}"
r += 1

# Swedish 2026 (partial)
ws.cell(row=r, column=1, value="Swedish 2026 partial (Jan-Apr 2026 only)")
ws.cell(row=r, column=2, value=f"=SUMIF(G{first_data_row}:G{last_data_row},\"2026\",E{first_data_row}:E{last_data_row})")
ws.cell(row=r, column=3, value="Will need more data later").font = It
r += 2

# Notes
ws.cell(row=r, column=1, value="Notes:").font = S
r += 1
ws.cell(row=r, column=1, value="• Each month has both a 'GROSS INTEREST TO <date>' line and (sometimes) an 'ADDED GROSS INTEREST' line. Both credit the account.").font = It
r += 1
ws.cell(row=r, column=1, value="• UK 25/26 PSA covers £1,000 for basic rate; with ~£240 interest, no UK tax due — still reportable on SA100 box 2.").font = It
r += 1
ws.cell(row=r, column=1, value="• Sweden taxes the credited interest at 30% capital rate (Box 7.2 on INK1).").font = It
r += 1
ws.cell(row=r, column=1, value="• MISSING rows (Jan-Apr 2025) need earlier statements to complete Swedish 2025 figure.").font = It
r += 1

# Column widths
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 16
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 24
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 12

wb.save(PATH)

# Compute totals for chat
uk_data = [d for d in data if d[4] == "25/26" and d[2] is not None]
uk_total = sum((d[2] or 0) + (d[3] or 0) for d in uk_data)
swed_2025 = [d for d in data if d[5] == "2025" and d[2] is not None]
swed_total = sum((d[2] or 0) + (d[3] or 0) for d in swed_2025)
print(f"UK 25/26 total interest (12 months ✓): £{uk_total:.2f}")
print(f"Swedish 2025 partial (May-Dec only, 4 months missing): £{swed_total:.2f}")
