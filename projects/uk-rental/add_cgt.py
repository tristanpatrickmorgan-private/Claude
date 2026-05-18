"""Add Capital Gains 2025 tab to the tax return spreadsheet."""
import openpyxl
from openpyxl.styles import Font, PatternFill

PATH = '/home/user/Claude/projects/uk-rental/Tax_Return.xlsx'
wb = openpyxl.load_workbook(PATH)

name = "Capital Gains 2025"
if name in wb.sheetnames:
    del wb[name]
ws = wb.create_sheet(name)

# Styles
H = Font(bold=True, size=14, color="FFFFFF")
HF = PatternFill("solid", fgColor="4472C4")
S = Font(bold=True, size=11)
SF = PatternFill("solid", fgColor="D9E1F2")
R = PatternFill("solid", fgColor="FFE699")
G = PatternFill("solid", fgColor="C6EFCE")
Bld = Font(bold=True)
It = Font(italic=True)

# Title
ws.cell(row=1, column=1, value="Capital Gains 2025 — Vanguard ISA (sold June 2025)").font = H
ws.cell(row=1, column=1).fill = HF
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

# Context
ws.cell(row=3, column=1, value="Held in UK ISA → no UK tax. Swedish residency 25 Nov 2021 → cost basis taken at that date.").font = It
ws.cell(row=4, column=1, value="Swedish capital gains tax: 30% flat on net gain in SEK.").font = It

# === Parameters (editable) ===
r = 6
ws.cell(row=r, column=1, value="Parameters (edit to recalc):").font = S
for c in range(1, 6): ws.cell(row=r, column=c).fill = SF
r += 1

ws.cell(row=r, column=1, value="Cost basis date").font = Bld
ws.cell(row=r, column=2, value="25 Nov 2021")
ws.cell(row=r, column=3, value="(Swedish tax-residency start)").font = It
r += 1

ws.cell(row=r, column=1, value="Cost basis value (GBP)")
ws.cell(row=r, column=2, value=23031.31)
cost_gbp_ref = f"$B${r}"
r += 1

ws.cell(row=r, column=1, value="Cost basis FX rate (Nov 2021)")
ws.cell(row=r, column=2, value=11.84728)
ws.cell(row=r, column=3, value="Riksbanken Nov 2021 monthly avg").font = It
cost_fx_ref = f"$B${r}"
r += 1

ws.cell(row=r, column=1, value="Sale date").font = Bld
ws.cell(row=r, column=2, value="5 Jun 2025")
r += 1

ws.cell(row=r, column=1, value="Sale proceeds (GBP)")
ws.cell(row=r, column=2, value=29862.97)
sale_gbp_ref = f"$B${r}"
r += 1

ws.cell(row=r, column=1, value="Sale FX rate (Jun 2025)")
ws.cell(row=r, column=2, value=12.9485)
ws.cell(row=r, column=3, value="Riksbanken Jun 2025 monthly avg").font = It
sale_fx_ref = f"$B${r}"
r += 1

ws.cell(row=r, column=1, value="Swedish capital tax rate")
ws.cell(row=r, column=2, value=0.30)
rate_ref = f"$B${r}"
r += 2

# === Calculation ===
ws.cell(row=r, column=1, value="Calculation").font = S
for c in range(1, 6): ws.cell(row=r, column=c).fill = SF
r += 1
ws.cell(row=r, column=2, value="GBP").font = Bld
ws.cell(row=r, column=3, value="SEK").font = Bld
r += 1

# Cost basis
ws.cell(row=r, column=1, value="Cost basis (omkostnadsbelopp)")
ws.cell(row=r, column=2, value=f"={cost_gbp_ref}")
ws.cell(row=r, column=3, value=f"={cost_gbp_ref}*{cost_fx_ref}")
r += 1

# Sale proceeds
ws.cell(row=r, column=1, value="Sale proceeds (försäljningspris)")
ws.cell(row=r, column=2, value=f"={sale_gbp_ref}")
ws.cell(row=r, column=3, value=f"={sale_gbp_ref}*{sale_fx_ref}")
r += 1

# Gain
ws.cell(row=r, column=1, value="Capital gain").font = Bld
ws.cell(row=r, column=2, value=f"={sale_gbp_ref}-{cost_gbp_ref}").font = Bld
ws.cell(row=r, column=3, value=f"={sale_gbp_ref}*{sale_fx_ref}-{cost_gbp_ref}*{cost_fx_ref}").font = Bld
ws.cell(row=r, column=2).fill = R
ws.cell(row=r, column=3).fill = R
gain_gbp_ref = f"$B${r}"
gain_sek_ref = f"$C${r}"
r += 2

# Decomposition (optional/explanatory)
ws.cell(row=r, column=1, value="Decomposition (memo):").font = It
r += 1
ws.cell(row=r, column=1, value="  Investment performance in GBP")
ws.cell(row=r, column=2, value=f"={sale_gbp_ref}-{cost_gbp_ref}")
r += 1
ws.cell(row=r, column=1, value="  FX uplift on cost basis (Nov21 → Jun25)")
ws.cell(row=r, column=3, value=f"={cost_gbp_ref}*({sale_fx_ref}-{cost_fx_ref})")
r += 1
ws.cell(row=r, column=1, value="  → Total SEK gain (matches above)").font = It
r += 2

# Tax
ws.cell(row=r, column=1, value="Swedish capital tax (30% of SEK gain)").font = Bld
ws.cell(row=r, column=2, value=f"={gain_sek_ref}*{rate_ref}/{sale_fx_ref}").font = Bld
ws.cell(row=r, column=3, value=f"={gain_sek_ref}*{rate_ref}").font = Bld
ws.cell(row=r, column=2).fill = G
ws.cell(row=r, column=3).fill = G
tax_sek_ref = f"$C${r}"
r += 3

# === How to report in the Swedish return ===
ws.cell(row=r, column=1, value="Reporting in your Swedish tax return:").font = S
for c in range(1, 6): ws.cell(row=r, column=c).fill = SF
r += 1
ws.cell(row=r, column=1, value="• File K4 (bilaga K4) — avsnitt A or C depending on whether marknadsnoterade aktier/fonder").font = It
r += 1
ws.cell(row=r, column=1, value="• K4 lists: antal/units, försäljningspris, omkostnadsbelopp, vinst").font = It
r += 1
ws.cell(row=r, column=1, value="• Net gain rolls into Box 7.4 (kapitalvinst marknadsnoterade aktier/fonder) on INK1").font = It
r += 1
ws.cell(row=r, column=1, value="• Box 7.2 catches general interest/dividends; CGT-specific lines are 7.4/7.5/7.6 depending on type").font = It
r += 2

# === Assumptions & caveats ===
ws.cell(row=r, column=1, value="Assumptions & caveats:").font = S
for c in range(1, 6): ws.cell(row=r, column=c).fill = SF
r += 1
ws.cell(row=r, column=1, value="• Cost basis taken at 25 Nov 2021 value (Swedish residency start). This is the simplest approach if").font = It
r += 1
ws.cell(row=r, column=1, value="  the bulk of the holding pre-dated the move. If significant contributions were made AFTER Nov 2021,").font = It
r += 1
ws.cell(row=r, column=1, value="  Swedish tax law applies the average-cost method (genomsnittsmetoden) — each contribution's own").font = It
r += 1
ws.cell(row=r, column=1, value="  cost basis (in SEK at the time) is weighted into the average. Confirm with accountant if you").font = It
r += 1
ws.cell(row=r, column=1, value="  contributed materially between Nov 2021 and Jun 2025.").font = It
r += 1
ws.cell(row=r, column=1, value="• FX rates used are Riksbanken monthly averages for Nov 2021 and Jun 2025.").font = It
r += 1
ws.cell(row=r, column=1, value="• ISA wrapper provides no Swedish tax benefit — Sweden taxes the gain regardless.").font = It
r += 1
ws.cell(row=r, column=1, value="• Capital losses elsewhere in the same year can offset this gain (currently none assumed).").font = It

# Column widths
ws.column_dimensions['A'].width = 60
ws.column_dimensions['B'].width = 16
ws.column_dimensions['C'].width = 16

wb.save(PATH)
print("Saved.")

# Verify expected values
print(f"\nExpected values:")
print(f"  Cost basis GBP: £23,031.31  → SEK {23031.31 * 11.84728:,.2f}")
print(f"  Sale value GBP: £29,862.97  → SEK {29862.97 * 12.9485:,.2f}")
print(f"  Gain GBP:       £{29862.97 - 23031.31:,.2f}")
print(f"  Gain SEK:       SEK {29862.97 * 12.9485 - 23031.31 * 11.84728:,.2f}")
print(f"  Tax @ 30%:      SEK {(29862.97 * 12.9485 - 23031.31 * 11.84728) * 0.30:,.2f}")
print(f"  Tax @ 30% GBP:  £{(29862.97 * 12.9485 - 23031.31 * 11.84728) * 0.30 / 12.9485:,.2f}")
