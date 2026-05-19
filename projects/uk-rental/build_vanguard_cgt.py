"""Rebuild Capital Gains 2025 tab — per-fund genomsnittsmetoden from actual Vanguard ISA history.

4 funds, ~46 buy transactions, verified Riksbanken monthly Medel FX rates.
Pre-residency 15 Nov 2021 FTSE 100 partial sale handled (excluded from Swedish gain).
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PATH = '/home/user/Claude/projects/uk-rental/Tax_Return.xlsx'
wb = openpyxl.load_workbook(PATH)

name = "Capital Gains 2025"
if name in wb.sheetnames:
    del wb[name]
ws = wb.create_sheet(name)

# Fonts and fills
TITLE = Font(bold=True, size=14, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="4472C4")
SECTION = Font(bold=True, size=11, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="5B9BD5")
SUB = Font(bold=True, size=10)
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
HDR = Font(bold=True, size=10)
HDR_FILL = PatternFill("solid", fgColor="E7E6E6")
RESULT = PatternFill("solid", fgColor="FFE699")
GREEN = PatternFill("solid", fgColor="C6EFCE")
EDIT = PatternFill("solid", fgColor="FFFFCC")
EXCLUDED = PatternFill("solid", fgColor="F2F2F2")
EXCLUDED_FONT = Font(italic=True, color="808080")
BOLD = Font(bold=True)
ITALIC = Font(italic=True)
TOTAL_FONT = Font(bold=True, size=10)
TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")

# Number formats
SEK_FMT = '#,##0'
GBP_FMT = '#,##0.00'
UNITS_FMT = '0.0000'
FX_FMT = '0.00000'
PCT_FMT = '0.00%'

# ============================================================
# Riksbanken monthly Medel GBP/SEK (verified 19 May 2026)
# ============================================================
fx_rates = [
    ("Sep 2020", 11.45956),
    ("Oct 2020", 11.47127),
    ("Nov 2020", 11.41785),
    ("Dec 2020", 11.23895),
    ("Jan 2021", 11.31239),
    ("Feb 2021", 11.55250),
    ("Mar 2021", 11.83967),
    ("Apr 2021", 11.74595),
    ("May 2021", 11.75710),
    ("Jun 2021", 11.77512),
    ("Jul 2021", 11.90743),
    ("Aug 2021", 11.97744),
    ("Sep 2021", 11.87028),
    ("Oct 2021", 11.86876),
    ("Nov 2021", 11.84728),
    ("Jun 2025", 12.94850),
]

# ============================================================
# Transaction data (from Vanguard PDFs, 19 May 2026)
# Format: (date, units, gbp, fx_month_label, swedish_relevant_for_sales)
# ============================================================
ftse100_buys = [
    ("2020-09-25", 0.9826, 100.00, "Sep 2020"),
    ("2020-09-29", 12.5970, 1254.41, "Sep 2020"),
    ("2020-11-02", 1.0334, 100.00, "Nov 2020"),
    ("2020-12-01", 0.9124, 100.00, "Dec 2020"),
    ("2021-01-04", 0.8808, 100.00, "Jan 2021"),
    ("2021-02-01", 3.5798, 400.00, "Feb 2021"),
    ("2021-02-03", 8.8976, 1000.00, "Feb 2021"),
    ("2021-02-04", 8.8976, 1000.00, "Feb 2021"),
    ("2021-03-01", 3.5005, 400.00, "Mar 2021"),
    ("2021-03-05", 4.3373, 500.00, "Mar 2021"),
    ("2021-04-01", 3.4025, 400.00, "Apr 2021"),
    ("2021-04-20", 6.2246, 750.00, "Apr 2021"),
    ("2021-05-04", 3.3190, 400.00, "May 2021"),
    ("2021-06-01", 3.2354, 400.00, "Jun 2021"),
    ("2021-07-01", 3.1929, 400.00, "Jul 2021"),
    ("2021-08-02", 3.2245, 400.00, "Aug 2021"),
    ("2021-08-29", 1.5910, 200.00, "Aug 2021"),
    ("2021-09-01", 3.1681, 400.00, "Sep 2021"),
    ("2021-10-01", 3.2146, 400.00, "Oct 2021"),
    ("2021-11-01", 3.0974, 400.00, "Nov 2021"),
]
ftse100_sales = [
    ("2021-11-15", 5.0134, 657.26, "Nov 2021", False),  # pre-residency
    ("2025-06-09", 41.7177, 7500.00, "Jun 2025", True),
    ("2025-06-17", 32.5579, 5861.72, "Jun 2025", True),
]

vusa_buys = [
    ("2020-09-25", 2, 98.25, "Sep 2020"),
    ("2020-09-29", 25, 1244.85, "Sep 2020"),
    ("2020-11-02", 2, 97.46, "Nov 2020"),
    ("2020-12-01", 1, 51.99, "Dec 2020"),
    ("2020-12-01", 1, 52.09, "Dec 2020"),
    ("2021-01-04", 1, 52.42, "Jan 2021"),
    ("2021-01-06", 1, 51.63, "Jan 2021"),
    ("2021-02-01", 7, 362.81, "Feb 2021"),
    ("2021-02-03", 18, 963.86, "Feb 2021"),
    ("2021-02-04", 18, 963.86, "Feb 2021"),
    ("2021-03-01", 7, 367.71, "Mar 2021"),
    ("2021-03-05", 9, 465.14, "Mar 2021"),
    ("2021-03-22", 7, 384.29, "Mar 2021"),
    ("2021-04-20", 13, 731.53, "Apr 2021"),
    ("2021-04-21", 6, 343.56, "Apr 2021"),
    ("2021-05-19", 7, 395.80, "May 2021"),
    ("2021-06-21", 6, 355.37, "Jun 2021"),
    ("2021-07-21", 6, 361.88, "Jul 2021"),
    ("2021-08-19", 6, 376.08, "Aug 2021"),
    ("2021-09-21", 6, 361.55, "Sep 2021"),
    ("2021-10-20", 6, 385.28, "Oct 2021"),
]
vusa_sales = [
    ("2025-06-09", 90, 7544.99, "Jun 2025", True),
    ("2025-06-17", 65, 5456.78, "Jun 2025", True),
]

vmid_buys = [
    ("2021-02-10", 30, 993.41, "Feb 2021"),
    ("2021-03-05", 7, 233.18, "Mar 2021"),
    ("2021-04-20", 8, 277.38, "Apr 2021"),
]
vmid_sales = [
    ("2025-06-05", 45, 1457.21, "Jun 2025", True),
]

verx_buys = [
    ("2021-02-10", 34, 974.39, "Feb 2021"),
    ("2021-03-14", 8, 232.16, "Mar 2021"),
]
verx_sales = [
    ("2025-06-05", 42, 1541.25, "Jun 2025", True),
]

# ============================================================
# Build sheet
# ============================================================

# Title row
ws.cell(row=1, column=1, value="Vanguard ISA — Capital Gains 2025 (genomsnittsmetoden, per fund)").font = TITLE
ws.cell(row=1, column=1).fill = TITLE_FILL
ws.cell(row=1, column=1).alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 22
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

# Methodology notes
notes = [
    "Per Swedish tax law (44 kap. 14 § IL): omkostnadsbelopp = original acquisition cost in SEK at original-date FX. No rebasing on residency move.",
    "Genomsnittsmetoden applied per security. 4 funds traded: FTSE 100 UT Acc, VUSA, VMID, VERX. Pre-residency partial sale 15 Nov 2021 (FTSE 100) excluded.",
    "FX = Riksbanken monthly Medel GBP/SEK (all verified 19 May 2026). Edit yellow cells to override.",
    "Sources: Vanguard ISA VG0208845 transaction history PDFs (19 May 2026); Riksbanken sok-rantor-och-valutakurser.",
]
for i, note in enumerate(notes):
    ws.cell(row=3 + i, column=1, value=note).font = ITALIC
    ws.merge_cells(start_row=3+i, start_column=1, end_row=3+i, end_column=6)

# ============================================================
# K4 Summary (top)
# ============================================================
r = 9
ws.cell(row=r, column=1, value="K4 avsnitt A — rows to enter on Skatteverket e-tjänst").font = SECTION
ws.cell(row=r, column=1).fill = SECTION_FILL
for c in range(1, 7):
    ws.cell(row=r, column=c).fill = SECTION_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
r += 1

# K4 headers
k4_headers = ["Antal", "Beteckning", "Försäljningspris (SEK)", "Omkostnadsbelopp (SEK)", "Vinst (SEK)"]
for i, h in enumerate(k4_headers, 1):
    ws.cell(row=r, column=i, value=h).font = HDR
    ws.cell(row=r, column=i).fill = HDR_FILL
k4_header_row = r
r += 1

# Placeholder K4 row positions — we'll fill formula refs after computing per-fund blocks
K4_FTSE_ROW = r; r += 1
K4_VUSA_ROW = r; r += 1
K4_VMID_ROW = r; r += 1
K4_VERX_ROW = r; r += 1
K4_TOTAL_ROW = r; r += 1

# Build placeholder cells (will fill formula refs later)
ws.cell(row=K4_FTSE_ROW, column=2, value="FTSE 100 Index Unit Trust Accumulation")
ws.cell(row=K4_VUSA_ROW, column=2, value="S&P 500 UCITS ETF — Distributing (VUSA)")
ws.cell(row=K4_VMID_ROW, column=2, value="FTSE 250 UCITS ETF — Distributing (VMID)")
ws.cell(row=K4_VERX_ROW, column=2, value="FTSE Developed Europe ex UK UCITS ETF — Distributing (VERX)")
ws.cell(row=K4_TOTAL_ROW, column=2, value="Total").font = TOTAL_FONT
for c in range(1, 6):
    ws.cell(row=K4_TOTAL_ROW, column=c).fill = TOTAL_FILL
ws.cell(row=K4_TOTAL_ROW, column=3, value=f"=SUM(C{K4_FTSE_ROW}:C{K4_VERX_ROW})").font = TOTAL_FONT
ws.cell(row=K4_TOTAL_ROW, column=4, value=f"=SUM(D{K4_FTSE_ROW}:D{K4_VERX_ROW})").font = TOTAL_FONT
ws.cell(row=K4_TOTAL_ROW, column=5, value=f"=SUM(E{K4_FTSE_ROW}:E{K4_VERX_ROW})").font = TOTAL_FONT
ws.cell(row=K4_TOTAL_ROW, column=5).fill = RESULT

r += 1

# ============================================================
# Tax calculation block
# ============================================================
ws.cell(row=r, column=1, value="Box 7.4 + Swedish tax").font = SECTION
ws.cell(row=r, column=1).fill = SECTION_FILL
for c in range(1, 7):
    ws.cell(row=r, column=c).fill = SECTION_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
r += 1

ws.cell(row=r, column=1, value="Box 7.4 — Vinst fondandelar m.m. (SEK)").font = BOLD
ws.cell(row=r, column=3, value=f"=E{K4_TOTAL_ROW}").font = BOLD
ws.cell(row=r, column=3).fill = RESULT
BOX_74_ROW = r
r += 1

ws.cell(row=r, column=1, value="Swedish capital tax rate")
ws.cell(row=r, column=3, value=0.30).number_format = PCT_FMT
TAX_RATE_ROW = r
r += 1

ws.cell(row=r, column=1, value="Tax on Vanguard gain (SEK)").font = BOLD
ws.cell(row=r, column=3, value=f"=C{BOX_74_ROW}*C{TAX_RATE_ROW}").font = BOLD
ws.cell(row=r, column=3).fill = GREEN
r += 2

# Format SEK cells in K4 + tax block
for rr in [K4_FTSE_ROW, K4_VUSA_ROW, K4_VMID_ROW, K4_VERX_ROW, K4_TOTAL_ROW]:
    ws.cell(row=rr, column=1).number_format = UNITS_FMT
    ws.cell(row=rr, column=3).number_format = SEK_FMT
    ws.cell(row=rr, column=4).number_format = SEK_FMT
    ws.cell(row=rr, column=5).number_format = SEK_FMT
ws.cell(row=BOX_74_ROW, column=3).number_format = SEK_FMT

# ============================================================
# FX rates block
# ============================================================
ws.cell(row=r, column=1, value="FX rates — Riksbanken monthly Medel GBP/SEK (verified)").font = SECTION
ws.cell(row=r, column=1).fill = SECTION_FILL
for c in range(1, 7):
    ws.cell(row=r, column=c).fill = SECTION_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
r += 1

ws.cell(row=r, column=1, value="Month").font = HDR
ws.cell(row=r, column=2, value="FX rate (SEK per £1)").font = HDR
ws.cell(row=r, column=3, value="Source").font = HDR
for c in range(1, 7):
    ws.cell(row=r, column=c).fill = HDR_FILL
r += 1

fx_row = {}
for month, rate in fx_rates:
    ws.cell(row=r, column=1, value=month)
    ws.cell(row=r, column=2, value=rate).number_format = FX_FMT
    ws.cell(row=r, column=2).fill = EDIT
    ws.cell(row=r, column=3, value="Riksbanken Medel")
    fx_row[month] = r
    r += 1

r += 1

# ============================================================
# Helper to build per-fund block
# ============================================================
def fx_ref(month):
    """Return absolute reference to FX rate cell for given month."""
    return f"$B${fx_row[month]}"

def build_fund_block(name, ticker, buys, sales, start_row, k4_target_row):
    """Build a per-fund block. Returns final row used."""
    r = start_row
    # Section header
    ws.cell(row=r, column=1, value=f"Fund: {name}").font = SECTION
    ws.cell(row=r, column=1).fill = SECTION_FILL
    for c in range(1, 7):
        ws.cell(row=r, column=c).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1

    # Buys subsection
    ws.cell(row=r, column=1, value="Buys").font = SUB
    ws.cell(row=r, column=1).fill = SUB_FILL
    for c in range(1, 7):
        ws.cell(row=r, column=c).fill = SUB_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1

    # Buy headers
    for i, h in enumerate(["Date", "Units", "GBP", "FX month", "FX rate", "SEK cost"], 1):
        ws.cell(row=r, column=i, value=h).font = HDR
        ws.cell(row=r, column=i).fill = HDR_FILL
    r += 1

    first_buy = r
    for date, units, gbp, fxmonth in buys:
        ws.cell(row=r, column=1, value=date)
        ws.cell(row=r, column=2, value=units).number_format = UNITS_FMT
        ws.cell(row=r, column=3, value=gbp).number_format = GBP_FMT
        ws.cell(row=r, column=4, value=fxmonth)
        ws.cell(row=r, column=5, value=f"={fx_ref(fxmonth)}").number_format = FX_FMT
        ws.cell(row=r, column=6, value=f"=C{r}*E{r}").number_format = SEK_FMT
        r += 1
    last_buy = r - 1

    # Total buys
    ws.cell(row=r, column=1, value="Total buys").font = TOTAL_FONT
    ws.cell(row=r, column=2, value=f"=SUM(B{first_buy}:B{last_buy})").number_format = UNITS_FMT
    ws.cell(row=r, column=3, value=f"=SUM(C{first_buy}:C{last_buy})").number_format = GBP_FMT
    ws.cell(row=r, column=6, value=f"=SUM(F{first_buy}:F{last_buy})").number_format = SEK_FMT
    for c in [2, 3, 6]:
        ws.cell(row=r, column=c).font = TOTAL_FONT
    for c in range(1, 7):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
    total_units_buy = f"$B${r}"
    total_sek_buy = f"$F${r}"
    r += 1

    # Avg cost
    ws.cell(row=r, column=1, value="Average cost per unit (SEK)").font = BOLD
    ws.cell(row=r, column=6, value=f"={total_sek_buy}/{total_units_buy}").number_format = SEK_FMT
    avg_cost_ref = f"$F${r}"
    r += 2

    # Sales subsection
    ws.cell(row=r, column=1, value="Sales").font = SUB
    ws.cell(row=r, column=1).fill = SUB_FILL
    for c in range(1, 7):
        ws.cell(row=r, column=c).fill = SUB_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1

    for i, h in enumerate(["Date", "Units", "GBP", "FX month", "FX rate", "SEK proceeds"], 1):
        ws.cell(row=r, column=i, value=h).font = HDR
        ws.cell(row=r, column=i).fill = HDR_FILL
    ws.cell(row=r, column=7 if False else 6, value="").font = HDR  # spacer
    # Add "Swedish?" col
    ws.cell(row=r, column=7, value="Swedish 2025?").font = HDR
    ws.cell(row=r, column=7).fill = HDR_FILL
    r += 1

    first_sale = r
    swedish_sale_rows = []
    for date, units, gbp, fxmonth, is_swedish in sales:
        ws.cell(row=r, column=1, value=date)
        ws.cell(row=r, column=2, value=units).number_format = UNITS_FMT
        ws.cell(row=r, column=3, value=gbp).number_format = GBP_FMT
        ws.cell(row=r, column=4, value=fxmonth)
        ws.cell(row=r, column=5, value=f"={fx_ref(fxmonth)}").number_format = FX_FMT
        ws.cell(row=r, column=6, value=f"=C{r}*E{r}").number_format = SEK_FMT
        ws.cell(row=r, column=7, value="Yes" if is_swedish else "No — pre-residency")
        if not is_swedish:
            for c in range(1, 8):
                ws.cell(row=r, column=c).fill = EXCLUDED
                ws.cell(row=r, column=c).font = EXCLUDED_FONT
        else:
            swedish_sale_rows.append(r)
        r += 1
    last_sale = r - 1

    # Swedish 2025 sale totals
    if swedish_sale_rows:
        swedish_units_formula = "+".join([f"B{rr}" for rr in swedish_sale_rows])
        swedish_proceeds_formula = "+".join([f"F{rr}" for rr in swedish_sale_rows])
        ws.cell(row=r, column=1, value="Swedish 2025 sale total").font = TOTAL_FONT
        ws.cell(row=r, column=2, value=f"={swedish_units_formula}").number_format = UNITS_FMT
        ws.cell(row=r, column=6, value=f"={swedish_proceeds_formula}").number_format = SEK_FMT
        for c in [2, 6]:
            ws.cell(row=r, column=c).font = TOTAL_FONT
        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = TOTAL_FILL
        swedish_units_ref = f"$B${r}"
        swedish_proceeds_ref = f"$F${r}"
        r += 2

    # K4 calc
    ws.cell(row=r, column=1, value="K4 calculation").font = SUB
    ws.cell(row=r, column=1).fill = SUB_FILL
    for c in range(1, 7):
        ws.cell(row=r, column=c).fill = SUB_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1

    ws.cell(row=r, column=1, value="Units sold in 2025 (Antal)")
    ws.cell(row=r, column=6, value=f"={swedish_units_ref}").number_format = UNITS_FMT
    units_2025_ref = f"$F${r}"
    r += 1

    ws.cell(row=r, column=1, value="Försäljningspris (SEK)")
    ws.cell(row=r, column=6, value=f"={swedish_proceeds_ref}").number_format = SEK_FMT
    proceeds_ref = f"$F${r}"
    r += 1

    ws.cell(row=r, column=1, value="Omkostnadsbelopp = units × avg cost (SEK)")
    ws.cell(row=r, column=6, value=f"={units_2025_ref}*{avg_cost_ref}").number_format = SEK_FMT
    cost_basis_ref = f"$F${r}"
    r += 1

    ws.cell(row=r, column=1, value="Vinst (SEK)").font = BOLD
    ws.cell(row=r, column=6, value=f"={proceeds_ref}-{cost_basis_ref}").number_format = SEK_FMT
    ws.cell(row=r, column=6).fill = RESULT
    ws.cell(row=r, column=6).font = BOLD
    gain_ref = f"$F${r}"
    r += 1

    # Wire up K4 summary row at top
    ws.cell(row=k4_target_row, column=1, value=f"={units_2025_ref}").number_format = UNITS_FMT
    ws.cell(row=k4_target_row, column=3, value=f"={proceeds_ref}").number_format = SEK_FMT
    ws.cell(row=k4_target_row, column=4, value=f"={cost_basis_ref}").number_format = SEK_FMT
    ws.cell(row=k4_target_row, column=5, value=f"={gain_ref}").number_format = SEK_FMT

    r += 2
    return r

# Build per-fund blocks
r = build_fund_block("FTSE 100 Index Unit Trust Accumulation", "FTSE100", ftse100_buys, ftse100_sales, r, K4_FTSE_ROW)
r = build_fund_block("S&P 500 UCITS ETF — Distributing (VUSA)", "VUSA", vusa_buys, vusa_sales, r, K4_VUSA_ROW)
r = build_fund_block("FTSE 250 UCITS ETF — Distributing (VMID)", "VMID", vmid_buys, vmid_sales, r, K4_VMID_ROW)
r = build_fund_block("FTSE Developed Europe ex UK UCITS ETF — Distributing (VERX)", "VERX", verx_buys, verx_sales, r, K4_VERX_ROW)

# ============================================================
# Excluded section
# ============================================================
ws.cell(row=r, column=1, value="Excluded — FTSE Developed Europe ex-U.K. Equity Index Fund (Accumulation)").font = SECTION
ws.cell(row=r, column=1).fill = SECTION_FILL
for c in range(1, 7):
    ws.cell(row=r, column=c).fill = SECTION_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
r += 1
ws.cell(row=r, column=1, value="Bought 05 Mar 2021 (£250, 0.8601 units) and sold 14 Mar 2021 (£259.10, 0.8601 units). Both pre-Swedish-residency. Not on Swedish return.").font = ITALIC
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
r += 2

# ============================================================
# Footer notes
# ============================================================
ws.cell(row=r, column=1, value="Method & references").font = SECTION
ws.cell(row=r, column=1).fill = SECTION_FILL
for c in range(1, 7):
    ws.cell(row=r, column=c).fill = SECTION_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
r += 1

footer_notes = [
    "• Genomsnittsmetoden applied per security (ISIN). Each fund has its own omkostnadsbelopp pool.",
    "• Each buy converted to SEK at Riksbanken monthly Medel GBP/SEK for the buy month. Sale proceeds converted at the sale month's Medel.",
    "• FTSE 100 had a pre-residency partial sale on 15 Nov 2021 (5.0134 units, £657.26). That sale is not on the Swedish return; it just reduces the unit pool. The remaining units carry the same average cost.",
    "• Swedish tax resident from 25 Nov 2021. All buys completed by 1 Nov 2021 — all pre-residency, but pre-residency buys still constitute the omkostnadsbelopp under Swedish rules (no rebasing).",
    "• ISA wrapper is irrelevant for Swedish tax — the full gain is taxable. Lossen från 2025 can offset (100% mot marknadsnoterade aktier/fonder; 70% mot annan kapital).",
    "• File K4 avsnitt A with the 4 rows above. Box 7.4 = sum of the 4 vinst figures.",
    "• Sources: 44 kap. 14 § Inkomstskattelagen; Skatteverket 'Utländsk valuta'; Vanguard ISA transaction PDFs (19 May 2026); Riksbanken sok-rantor-och-valutakurser.",
]
for note in footer_notes:
    ws.cell(row=r, column=1, value=note).font = ITALIC
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1

# Column widths
ws.column_dimensions['A'].width = 42
ws.column_dimensions['B'].width = 13
ws.column_dimensions['C'].width = 13
ws.column_dimensions['D'].width = 11
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 14
ws.column_dimensions['G'].width = 18

# Freeze panes below K4 summary
ws.freeze_panes = "A9"

# ============================================================
# Update Tax 25-26 sheet to point to new gain cell
# ============================================================
wb.save(PATH)

# Find new Box 7.4 row and patch Tax 25-26
wb2 = openpyxl.load_workbook(PATH)
ws2 = wb2['Capital Gains 2025']
box_74_row = None
for rr in range(1, ws2.max_row + 1):
    v = ws2.cell(row=rr, column=1).value
    if v and "Box 7.4" in str(v):
        box_74_row = rr
        break

if box_74_row:
    ws_tax = wb2['Tax 25-26']
    # Find row referencing Capital Gains 2025
    for rr in range(1, ws_tax.max_row + 1):
        v = ws_tax.cell(row=rr, column=2).value
        if v and isinstance(v, str) and "Capital Gains 2025" in v:
            ws_tax.cell(row=rr, column=2, value=f"='Capital Gains 2025'!$C${box_74_row}")
            # if col 3 also pointed, leave or null
            if ws_tax.cell(row=rr, column=3).value and "Capital Gains 2025" in str(ws_tax.cell(row=rr, column=3).value):
                ws_tax.cell(row=rr, column=3, value=f"='Capital Gains 2025'!$C${box_74_row}")
            print(f"Updated Tax 25-26 row {rr} to point to Capital Gains 2025!C{box_74_row}")
            break

wb2.save(PATH)

# ============================================================
# Compute & print preview
# ============================================================
fx = dict(fx_rates)

def fund_calc(buys, sales):
    total_units = sum(b[1] for b in buys)
    total_sek_cost = sum(b[1] * fx[b[3]] / b[1] * b[2] * fx[b[3]] / b[2] if False else b[2] * fx[b[3]] for b in buys)
    # Simpler: SEK cost = sum of GBP * FX per buy
    total_sek_cost = sum(b[2] * fx[b[3]] for b in buys)
    avg = total_sek_cost / total_units

    swedish_units = sum(s[1] for s in sales if s[4])
    swedish_gbp = sum(s[2] for s in sales if s[4])
    swedish_proceeds_sek = sum(s[2] * fx[s[3]] for s in sales if s[4])
    cost_basis_sek = swedish_units * avg
    gain = swedish_proceeds_sek - cost_basis_sek
    return total_units, total_sek_cost, avg, swedish_units, swedish_proceeds_sek, cost_basis_sek, gain

print("\n=== K4 Summary ===")
total_gain = 0
for label, buys, sales in [
    ("FTSE 100", ftse100_buys, ftse100_sales),
    ("VUSA", vusa_buys, vusa_sales),
    ("VMID", vmid_buys, vmid_sales),
    ("VERX", verx_buys, verx_sales),
]:
    tu, tc, avg, su, sp, cb, g = fund_calc(buys, sales)
    total_gain += g
    print(f"{label:10s}: units {su:>9.4f}  försälj {sp:>10,.0f}  omkostn {cb:>10,.0f}  vinst {g:>10,.0f}")
print(f"{'Total':10s}: {'':>9}             {'':>10}              {'':>10}        {total_gain:>10,.0f}")
print(f"\nBox 7.4: SEK {total_gain:,.0f}")
print(f"Tax @ 30%: SEK {total_gain * 0.30:,.0f}")
print(f"\nSaved: {PATH}")
